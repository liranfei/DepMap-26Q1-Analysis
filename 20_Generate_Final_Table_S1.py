import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
import os

# ===自动获取桌面路径 ===
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
final_path = os.path.join(desktop, "final_targets.csv")
sig_all_path = os.path.join(desktop, "significant_pairs.csv")

# 检查文件是否存在
if not os.path.exists(final_path) or not os.path.exists(sig_all_path):
    print("错误：桌面上缺少 final_targets.csv 或 significant_pairs.csv")
else:
    final = pd.read_csv(final_path)
    sig_all = pd.read_csv(sig_all_path)

    # === 1. 整理数据 ===
    final = final.sort_values(["Cancer", "Selectivity"])

    # 拆分基因名和Entrez ID
    final[["Gene_Symbol", "Entrez_ID"]] = final["Gene"].str.extract(r'^(.+?)\s*\((\d+)\)$')

    # 格式化数值 (保留4位小数，科学计数法表示p值)
    final["Chronos_median"] = final["Chronos_median"].round(4)
    final["Selectivity"]    = final["Selectivity"].round(4)
    final["p_value"]        = final["p_value"].apply(lambda x: f"{x:.3e}")
    final["q_value"]        = final["q_value"].apply(lambda x: f"{x:.3e}")

    # 筛选并重命名列
    table_s1 = final[[
        "Gene_Symbol", "Entrez_ID", "Cancer",
        "Chronos_median", "Selectivity",
        "n_target", "p_value", "q_value", "confidence"
    ]].copy()

    table_s1.columns = [
        "Gene Symbol", "Entrez ID", "Cancer Lineage",
        "Chronos Median", "Selectivity Score",
        "n (target lineage)", "p-value (Wilcoxon)",
        "q-value (BH-FDR)", "Confidence Score"
    ]

    # === 2. 创建 Excel 工作簿 ===
    wb = Workbook()

    # ---- Sheet 1: Table S1 主表 ----
    ws1 = wb.active
    ws1.title = "Table S1 - Final Targets"

    # 统一样式定义
    header_fill = PatternFill("solid", fgColor="1F3864") # 深蓝色标题
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # 写标题行
    for col_idx, col_name in enumerate(table_s1.columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Confidence 颜色映射
    conf_colors = {"High": "E8F5E9", "Medium": "FFF9C4", "Low": "FFEBEE"}

    # 写数据行
    for row_idx, row in enumerate(table_s1.itertuples(index=False), 2):
        conf = row[-1] # 取最后一列 Confidence Score
        row_fill = PatternFill("solid", fgColor=conf_colors.get(conf, "FFFFFF"))
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = center_align

    # 自动冻结首行
    ws1.freeze_panes = "A2"

    # ---- Sheet 2: 统计摘要 (Summary Statistics) ----
    ws2 = wb.create_sheet("Summary Statistics")
    summary_data = {
        "Metric": [
            "Total cell lines analyzed", "Genome-wide genes screened", 
            "High-variance genes (Top 500)", "Initial significant pairs (FDR < 0.05)",
            "Final high-confidence pairs (Selectivity < -0.5)", "Unique target genes (final)",
            "Cancer lineages represented", "High confidence (n>=15)", "Permutation test Z-score"
        ],
        "Value": [
            1208, 18532, 500, 274, 93, 57, 20, 
            int((final["confidence"] == "High").sum()), 75.7
        ]
    }
    
    # 写入摘要
    for r_idx, (m, v) in enumerate(zip(summary_data["Metric"], summary_data["Value"]), 1):
        ws2.cell(row=r_idx, column=1, value=m).font = Font(bold=True)
        ws2.cell(row=r_idx, column=2, value=v).alignment = center_align

    # ---- 保存文件 ----
    output_path = os.path.join(desktop, "Table_S1_Final_Targets.xlsx")
    wb.save(output_path)

    print(f"---  Table S1 制作完成 ---")
    print(f"保存路径: {output_path}")
    print(f"包含数据: {len(table_s1)} 条高置信度依赖项")
